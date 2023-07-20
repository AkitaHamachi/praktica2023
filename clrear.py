import pandas as pd
import requests
from bs4 import BeautifulSoup as bf, BeautifulSoup
import urllib.request
from openpyxl import load_workbook
URL_TEMPLATE = "https://catalog.data.gov/dataset/drug-overdose-death-rates-by-drug-type-sex-age-race-and-hispanic-origin-united-states-3f72f"
r = requests.get(URL_TEMPLATE)
soup = BeautifulSoup(r.text, "html.parser")
sum=0
fn='2.xlsx'
wb=load_workbook('2.xlsx')
ws=wb['Лист1']
for i in range(2, 10000):
        for cellObj in ws[f'A{i}':f'AZ{i}']:
            for cell in cellObj:
                ws[cell.coordinate] = ''
i=0
l=['Keyword usage','Categories','Geo search','Time based search','AccessURL accessibility','DownloadURL','DownloadURL accessibility','Rights','File size','Date of issue','Modification date','sum','row','col']
for col in range(1,14+1):
    cell=ws.cell(row=1, column=col)
    cell.value=l[i]
    i+=1
#ws.append(['Keyword usage','Categories','Geo search','Time based search','AccessURL accessibility','DownloadURL','DownloadURL accessibility','Rights','File size','Date of issue','Modification date','sum','row','col'])
if(r.status_code==200):
    sum+=50
    print('ссылка на метаданные(AccessURL accessibility 50)')
    ws['E2']='+'
else:
    print('ссылки на метаданные нет(AccessURL accessibility 50)')
    ws['E2'] = '-'
href=soup.findAll('a', class_='btn btn-primary'  )
tmphrf=""
for link in soup.find_all('a', class_='btn btn-primary'):
    if(link.get('href').find(".csv")!=-1):
        #print(link.get('href'))
        sum+=20
        print('есть ссылка на датасет(DownloadURL 20)')
        ws['F2'] = '+'
        tmphrf=link.get('href')
        break
if(tmphrf==""):
    print('нет ссылки на датасет(DownloadURL 20)')
    ws['F2'] = '-'
if(requests.get(tmphrf).status_code==200):
    urllib.request.urlretrieve(tmphrf, '1.csv')
    sum+=30
    print('ссылка на датасет работает(DownloadURL accessibility 30)')
    ws['G2'] = '+'
else:
    print('ссылка на датасет не работает(DownloadURL accessibility 30)')
    ws['G2'] = '-'
category=soup.findAll('th', class_='dataset-label' ,string= 'Category' )
if(category!=[]):
    sum+=30
    print('есть категория (Categories 30)')
    ws['B2'] = '+'
else:
    print('нет категории (Categories 30)')
    ws['B2'] = '-'
geo = soup.findAll('div',class_='dataset-map')
geo1 = soup.findAll('th', class_='dataset-label' ,string= 'Spatial')
if(((geo) or (geo1))!=[]):
    sum+=20
    print('есть геоданные(Geo search 20)')
    ws['C2'] = '+'
else:
    print('нет карты(Geo search 20)')
    ws['C2'] = '-'
teg = soup.findAll('a', class_='tag')
if(teg!=[]):
    sum+=30
    print('есть минимум 1 ключевое слово(Keyword usage 30)')
    ws['A2'] = '+'
else:
    print('нет ключевых слов (Keyword usage 30)')
    ws['A2'] = '-'
pub=soup.findAll('th', class_='dataset-label' ,string= 'Data First Published' )
if(pub!=[]):
    sum+=5
    print('есть дата публикации(Date of issue 5)')
    ws['J2'] = '+'
else:
    print('нет даты публикации (Date of issue 5)')
    ws['J2'] = '-'
mod=soup.findAll('th', class_='dataset-label' ,string= 'Data Last Modified' )
if(mod!=[]):
    sum+=5
    print('есть дата модификации (Modification date 5)')
    ws['K2'] = '+'
else:
    print('нет даты модификации(Modification date 5)')
    ws['K2'] = '-'
rights=soup.findAll('th', class_='dataset-label' ,string= 'Rights' )
if(rights!=[]):
    sum+=5
    print('есть права (Rights 5)')
    ws['H2'] = '+'
else:
    print('нет прав(Rights 5)')
    ws['H2'] = '-'
temp=soup.findAll('th', class_='dataset-label' ,string= 'Temporal' )
if(temp!=[]):
    sum+=20
    print('есть время (Time based search(20))')
    ws['D2'] = '+'
else:
    print('нет времени для поиска Time based search(20)')
    ws['D2'] = '-'
print('нет размера файла(File size	5)')
ws['I2'] = '-'
print('оценка')
print(sum)
ws['L2'] = sum
df = pd.read_csv('1.csv')
print('датасет')
print(df)
#print('строки')
num_rows = df.shape[0]
ws['M2'] = num_rows
#print('столбцы')
num_stolb = df.shape[1]
ws['N2'] = num_stolb
#print(num_rows)
#print(num_stolb)
tmp_df=df.select_dtypes(include=['float64', 'int64'])
tmplist=list(tmp_df)
list_=list(df)
i=0

for col in range(7,num_stolb+7):
    cell=ws.cell(row=7+i, column=1)
    cell.value=list_[i]
    i+=1
ws['B6'] = 'пустые строки'
ws['C6'] = 'уникальные строки'
ws['D6'] = 'число 0'
ws['E6'] = 'максимум'
ws['F6'] = 'минимум'
ws['G6'] = 'среднее'

print('пустые строки')
print(df.isnull().sum())
missing_values = list(df.isnull().sum())
i=0
for col in range(7,num_stolb+7):
    cell=ws.cell(row=7+i, column=2)
    cell.value=missing_values[i]
    i+=1
print('уникальные строки')
print (df.nunique ())
unique_values = list(df.nunique())
i=0
for col in range(7,num_stolb+7):
    cell=ws.cell(row=7+i, column=3)
    cell.value=unique_values[i]
    i+=1

amount_zero = []
min_values = []
max_values = []
mean_values = []
for i in range(df.shape[1]):
    if(list_[i] in tmplist):
        column = df[list_[i]]
        count = column[column == 0].count()
        amount_zero.append(count)
        print('количество значений 0 в столбце ',list_[i],':', count)
    else:
        amount_zero.append("NaN")
i=0
for col in range(7,num_stolb+7):
    cell=ws.cell(row=7+i, column=4)
    cell.value=amount_zero[i]

    i+=1
print('////////////////////////////////////')
for i in range(df.shape[1]):
    if list_[i] in tmplist:
        column = df[list_[i]]
        count = column.min()
        min_values.append(count)
        print('Минимальное значений в столбце ',list_[i],':', count)
    else:
        min_values.append("NaN")
i=0
for col in range(7,num_stolb+7):
    cell=ws.cell(row=7+i, column=5)
    cell.value=min_values[i]
    i+=1
print('////////////////////////////////////')
for i in range(df.shape[1]):
    if list_[i] in tmplist:
        column = df[list_[i]]
        count = column.max()
        max_values.append(count)
        print('Максимальное значений в столбце ',list_[i],':', count)
    else:
        max_values.append("NaN")
i=0
for col in range(7,num_stolb+7):
    cell=ws.cell(row=7+i, column=6)
    cell.value=max_values[i]
    i+=1
print('////////////////////////////////////')
for i in range(df.shape[1]):
    if list_[i] in tmplist:
        column = df[list_[i]]
        count = column.mean()
        mean_values.append(count)
        print('Среднее значений в столбца ',list_[i],':', count)
    else:
        mean_values.append("NaN")
i=0
for col in range(7,num_stolb+7):
    cell=ws.cell(row=7+i, column=7)
    cell.value=mean_values[i]
    i+=1
print('////////////////////////////////////')
cell=ws.cell(row=9+num_stolb, column=1)
cell.value='ссылка'
cell=ws.cell(row=9+num_stolb, column=2)
cell.value = URL_TEMPLATE
wb.save(fn)
wb.close()
